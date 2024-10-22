import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def int_to_roman(num):
    if not 1 <= num <= 3999:
        raise ValueError("Number out of range (must be between 1 and 3999)")
    
    roman_numerals = [
        (1000, 'M'), (900, 'CM'), (500, 'D'), (400, 'CD'),
        (100, 'C'), (90, 'XC'), (50, 'L'), (40, 'XL'),
        (10, 'X'), (9, 'IX'), (5, 'V'), (4, 'IV'), (1, 'I')
    ]
    
    result = []
    for (value, symbol) in roman_numerals:
        while num >= value:
            result.append(symbol)
            num -= value
    return ''.join(result)

def dict_to_latex_table(data_dict, numbering=False):
    latex_code = []
    latex_code.append("\\begin{table}[h!]")
    latex_code.append("\\centering")
    latex_code.append("\\begin{tabular}{%s}" % ('c c c' if numbering else 'c c'))
    latex_code.append("\\toprule")
    
    # Header row
    header = []
    if numbering:
        header.append("\\textbf{No.}")
    header.extend(["\\textbf{Column 1}", "\\textbf{Column 2}"])
    latex_code.append(" & ".join(header) + " \\\\")
    latex_code.append("\\midrule")
    
    # Data rows
    for i, (key, value) in enumerate(data_dict.items(), start=1):
        row = []
        if numbering:
            row.append(int_to_roman(i))
        row.extend([str(key), str(value)])
        latex_code.append(" & ".join(row) + " \\\\")
    
    latex_code.append("\\bottomrule")
    latex_code.append("\\end{tabular}")
    latex_code.append("\\end{table}")
    
    return "\n".join(latex_code)

def dict_to_markdown_table(data_dict, numbering=False):
    markdown_code = []
    
    # Header row
    header = []
    if numbering:
        header.append("**No.**")
    header.extend(["**Column 1**", "**Column 2**"])
    markdown_code.append("| " + " | ".join(header) + " |")
    
    # Divider row
    divider = ["---"] * (3 if numbering else 2)
    markdown_code.append("| " + " | ".join(divider) + " |")
    
    # Data rows
    for i, (key, value) in enumerate(data_dict.items(), start=1):
        row = []
        if numbering:
            row.append(int_to_roman(i))
        row.extend([str(key), str(value)])
        markdown_code.append("| " + " | ".join(row) + " |")
    
    return "\n".join(markdown_code)

def dict_to_dataframe(data_dict, column_names=["Column 1", "Column 2"], remove_latex_artefacts=True):
    def remove_latex(text):
        return re.sub(r'\$\$.*?\$\$', '', text)
    
    # Create the DataFrame from the dictionary
    df = pd.DataFrame(list(data_dict.items()), columns=column_names)
    
    # Remove LaTeX artifacts if the parameter is set to True
    if remove_latex_artefacts:
        df = df.applymap(lambda x: remove_latex(str(x)) if isinstance(x, str) else x)
    
    return df




def df_to_excel(df, file_name, title=None):
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Define styles
    header_font = Font(name='Consolas', bold=True)
    cell_font = Font(name='Consolas')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    thick_border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
    thin_bottom_border = Border(bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust row height
    for row in ws.iter_rows():
        for cell in row:
            cell.font = cell_font
            cell.fill = white_fill
            cell.border = thin_bottom_border
            ws.row_dimensions[cell.row].height = 21
    
    # Add title if provided
    start_row = 1
    if title:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(df.columns))
        title_cell = ws.cell(row=start_row, column=1)
        title_cell.value = title
        title_cell.font = Font(name='Consolas', bold=True, size=14)
        title_cell.alignment = alignment
        start_row += 1

    # Add header
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = white_fill
        cell.border = thick_border
        cell.alignment = alignment
        ws.row_dimensions[start_row].height = 21
        
    for col in ws.columns:
        max_length = 25
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length
        
    # Add data rows
    for row_num, row_data in enumerate(df.values, start=start_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.font = cell_font
            cell.fill = white_fill
            cell.border = thin_bottom_border
            cell.alignment = alignment
        ws.row_dimensions[row_num].height = 21

    # Save the workbook
    wb.save(file_name)

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def write_dataframes_to_excel(dataframes, file_name):
    # Initialize a new workbook
    wb = Workbook()
    
    # Define styles
    consolas_font = Font(name='Consolas', size=11)
    bold_consolas_font = Font(name='Consolas', size=11, bold=True)
    
    # Border styles
    thin_border = Border(bottom=Side(style='thin'))
    thick_bottom_border = Border(bottom=Side(style='thick'))
    
    # Background fill (white)
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    
    for (name, df) in dataframes:
        # Create a new sheet with the given name
        ws = wb.create_sheet(title=name)
        
        # Convert the Polars dataframe to a Pandas dataframe
        pdf = df.to_pandas()

        # Write the dataframe to the sheet, row by row
        for r_idx, row in enumerate(dataframe_to_rows(pdf, index=False, header=True)):
            ws.append(row)
            # Set row height
            ws.row_dimensions[r_idx + 1].height = 21

            # Style the header row
            if r_idx == 0:
                for c_idx, cell in enumerate(ws[r_idx + 1], 1):
                    cell.font = bold_consolas_font
                    cell.border = thick_bottom_border
                    cell.fill = white_fill
            else:
                # Style the data rows
                for c_idx, cell in enumerate(ws[r_idx + 1], 1):
                    cell.font = consolas_font
                    cell.border = thin_border
                    cell.fill = white_fill
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = max_length + 2
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = adjusted_width

    # Remove the default sheet created by Workbook
    del wb['Sheet']
    
    # Save the workbook
    wb.save(file_name)

dataframes = [("Sheet1", pl.DataFrame({'col1': [1, 2], 'col2': [3, 4]})),
              ("Sheet2", pl.DataFrame({'col1': [5, 6], 'col2': [7, 8]}))]
write_dataframes_to_excel(dataframes, 'output.xlsx')
